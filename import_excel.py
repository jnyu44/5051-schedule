"""
import_excel.py — One-time script to seed the database from the Excel schedule.

Usage:
    py import_excel.py

Reads '3.12.26 - 3 week look ahead.xlsx', creates Week + Row + Cell records,
and populates the schedule app database.
"""
from datetime import date

import openpyxl

# Import app context so SQLAlchemy can connect
from app import app
from models import db, Week, Column as Col, Row, Cell

EXCEL_PATH = r"D:\Pondview Seattle Dropbox\Pondview Property Management\Technology\Lorium\Projects\5051\Schedule\3.12.26 - 3 week look ahead.xlsx"


def import_schedule():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    # Get column IDs (must already exist from seed)
    columns = {c.name.lower(): c.id for c in Col.query.all()}

    # Mapping: Excel columns by index → our column names
    # Based on the Excel structure:
    #   A = Contractor, B = Date, C = Task, D = Activities, E = Confirmed, F = Pending
    col_map = {
        0: "contractor",
        1: "date",
        2: "task",
        3: "activities",
        4: "confirmed",
        5: "pending",
    }

    current_week = None
    row_order = 0

    for excel_row in ws.iter_rows(min_row=7, max_row=ws.max_row, max_col=6, values_only=False):
        values = [cell.value for cell in excel_row]

        # Check if this is a week header row (e.g. "Next Week: 3/1/26")
        first_val = str(values[0] or "").strip()
        is_week_header = any(
            kw in first_val.lower()
            for kw in ["week:", "week ", "prerequisites"]
        ) and all(v is None for v in values[2:])

        if is_week_header:
            if "prerequisite" in first_val.lower():
                continue  # Skip the prerequisites row

            # Parse start date from label if possible
            start_date = None
            if values[1] and hasattr(values[1], "date"):
                start_date = values[1].date() if hasattr(values[1], "date") else None

            max_order = db.session.query(db.func.max(Week.sort_order)).scalar() or 0
            current_week = Week(
                label=first_val,
                start_date=start_date,
                sort_order=max_order + 1,
            )
            db.session.add(current_week)
            db.session.flush()
            row_order = 0
            continue

        # Skip header rows and blank rows
        if current_week is None:
            continue
        if all(v is None for v in values):
            continue
        # Skip column header rows like "Date", "TASK", "ACTIVITIES"
        if first_val.lower() in ("", "none") and str(values[2] or "").strip().upper() == "TASK":
            continue

        # It's a data row — create Row + Cells
        row_order += 1
        new_row = Row(week_id=current_week.id, sort_order=row_order)
        db.session.add(new_row)
        db.session.flush()

        for idx, col_name in col_map.items():
            val = values[idx]
            if val is None:
                val_str = ""
            elif hasattr(val, "strftime"):
                try:
                    val_str = f"{val.month}/{val.day}"
                except Exception:
                    val_str = str(val)
            else:
                val_str = str(val).strip()

            col_id = columns.get(col_name)
            if col_id:
                db.session.add(Cell(row_id=new_row.id, column_id=col_id, value=val_str))

    db.session.commit()
    print("[OK] Excel data imported successfully!")

    # Print summary
    weeks = Week.query.count()
    rows = Row.query.count()
    print(f"   {weeks} weeks, {rows} rows imported.")


if __name__ == "__main__":
    with app.app_context():
        # Only import if no weeks exist yet
        if Week.query.count() > 0:
            print("[WARN] Schedule already has data. Delete schedule.db to re-import.")
        else:
            import_schedule()
